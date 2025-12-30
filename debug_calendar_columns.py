import os
import openpyxl

def inspect_calendar():
    db_file = os.path.join('DB', 'CALENDARIO&CLASSIFICA.xlsx')
    if not os.path.exists(db_file):
        print("File non trovato")
        return

    try:
        wb = openpyxl.load_workbook(db_file, data_only=True)
        sheet_name = next((s for s in wb.sheetnames if 'CALENDARIO' in s.upper()), wb.sheetnames[0])
        print(f"SHEET: {sheet_name}")
        ws = wb[sheet_name]

        # Salta header (riga 1-4) e prendi riga 5 (esempio dati)
        rows = list(ws.iter_rows(min_row=5, max_row=5, values_only=True))
        if rows:
            row = rows[0]
            for idx, val in enumerate(row):
                print(f"COL {idx}: {val}")

    except Exception as e:
        print(e)

if __name__ == "__main__":
    inspect_calendar()
